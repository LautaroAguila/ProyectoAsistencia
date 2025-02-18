from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.views.generic import *
from inicio.models import Alumno, Escuela
from django.urls import reverse_lazy
from django.shortcuts import render, get_object_or_404
import openpyxl

class CrearAlumno(CreateView):
    model = Alumno
    template_name = "inicio/alumno/crear_alumno.html"
    success_url = reverse_lazy('inicio:listado_alumnos')
    fields = ['nombre', 'apellido', 'escuela', 'dias', 'turno_profesor']

class ListadoAlumnos(ListView):
    model = Alumno
    template_name = "inicio/alumno/listado_alumnos.html"
    context_object_name = 'alumnos'

    def get_queryset(self):
        # Obtener el turno guardado en la sesión
        turno_profesor = self.request.session.get('turno_profesor')

        # Si no hay turno seleccionado, no mostrar alumnos
        if not turno_profesor:
            return Alumno.objects.none()

        # Filtrar los alumnos según el turno del profesor
        queryset = Alumno.objects.filter(turno_profesor=turno_profesor)

        # Obtener el parámetro de ordenamiento
        orden = self.request.GET.get('orden', 'id')  # Por defecto se ordena por ID

        # Validar las opciones de ordenamiento permitidas
        opciones_validas = ['escuela', '-escuela', 'apellido', '-apellido']
        if orden not in opciones_validas:
            orden = 'id'

        # Retornar el conjunto de datos filtrado y ordenado
        return queryset.order_by(orden)

class VerAlumno(DetailView):
    model = Alumno
    template_name = "inicio/alumno/ver_alumno.html"

class EditarAlumno(UpdateView):
    model = Alumno
    template_name = "inicio/alumno/editar_alumno.html"
    success_url = reverse_lazy('inicio:listado_alumnos')
    fields = ['nombre','apellido','grado','edad','turno','dni','grupo_familiar','tel_contacto','nombre_tutor','dni_tutor','escuela','dias','asistencias','inasistencias','profesionales','observaciones']

class EliminarAlumno(DeleteView):
    model = Alumno
    template_name = "inicio/alumno/eliminar_alumno.html"
    success_url = reverse_lazy('inicio:listado_alumnos')

def ver_escuela(request, escuela_numero): 
    try:
        # Filtrar la escuela basada en el campo 'numero'
        escuela = get_object_or_404(Escuela, numero=escuela_numero)
        # Renderizar la plantilla si se encuentra la escuela
        return render(request, 'inicio/escuela/ver_escuela.html', {'escuela': escuela})
    except:
        # Renderizar una plantilla de error si no se encuentra
        return render(request, 'inicio/escuela/escuela_no_creada.html', {'escuela_id': escuela_numero})

class CrearEscuela(CreateView):
    model = Escuela
    template_name = "inicio/escuela/crear_escuela.html"
    success_url = reverse_lazy('inicio:listado_alumnos')
    fields = ['nombre_escuela','numero','direccion','tel_escuela' ,'mail_escuela' ,'nombre_directivo']

def vaciar_escuelas(request):
    if request.method == 'POST':
        Escuela.objects.all().delete()  # Vaciar todas las escuelas
        return redirect('inicio:listado_alumnos')  # Redirige a la lista de alumnos o a otra página

    return render(request, 'inicio/escuela/vaciar_escuelas.html')

def seleccionar_turno(request):
    if request.method == 'POST':
        # Obtener el turno enviado desde el formulario
        turno = request.POST.get('turno')

        # Validar si el turno es válido (T o M)
        if turno in ['T', 'M']:
            request.session['turno_profesor'] = turno  # Guardar el turno en la sesión
            return redirect('inicio:listado_alumnos')  # Redirigir al listado de alumnos

    # Renderizar el formulario de selección
    return render(request, 'inicio/seleccionar_turno.html')

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def export_alumnos_excel(request):
    """Genera un Excel con los alumnos según el turno seleccionado."""
    turno = request.GET.get('turno', 'M')  # Valor predeterminado: Mañana

    if turno not in ['M', 'T']:
        return HttpResponse("Turno inválido", status=400)

    alumnos = Alumno.objects.filter(turno_profesor=turno)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="alumnos_{turno}.xlsx"'

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f'Alumnos - Turno {turno}'

    # Encabezados
    headers = [
        'Nombre', 'Apellido', 'Edad', 'Escuela', 'Grado', 'Turno', 'DNI',
        'Grupo Familiar', 'Tel Contacto', 'Nombre Tutor', 'DNI Tutor',
        'Días Totales', 'Asistencias', 'Inasistencias', 'Porcentaje Asistencia',
        'Profesionales', 'Observaciones'
    ]
    sheet.append(headers)

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border_style = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                          top=Side(border_style="thin"), bottom=Side(border_style="thin"))

    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style

    # Agregar los alumnos filtrados
    for alumno in alumnos:
        sheet.append([
            alumno.nombre, alumno.apellido, alumno.edad, alumno.escuela, alumno.grado, alumno.turno, alumno.dni,
            alumno.grupo_familiar, alumno.tel_contacto, alumno.nombre_tutor, alumno.dni_tutor,
            alumno.dias, alumno.asistencias, alumno.inasistencias, f"{alumno.porcentaje:.1f}%",
            alumno.profesionales, alumno.observaciones
        ])

    # Ajuste de ancho automático
    for col in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        sheet.column_dimensions[col[0].column_letter].width = max_length + 2

    workbook.save(response)
    return response
